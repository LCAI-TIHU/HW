#!/nfs/tools/freeware/python/Python-3.7.5/bin/python3

################################################
## Function Notes   :
##     Gen RESET Matrix Defintion
################################################
# Release History   :           
# Version  Date        Author     Description
# 0.1      2019-12-31  wuyan03
################################################

import sys, re, os
import pprint
from openpyxl import Workbook
from openpyxl import load_workbook
#打开一个sheet
wb = load_workbook('reset_template.xlsx')
#获取一个sheet
#ws = wb.active
ws = wb.get_sheet_by_name('Sheet1')

#复制一个表格
ws1 = wb.create_sheet("My_sheet")  
#ws1 = wb.copy_worksheet(ws)

#define output txt file name
#filename = 'reset_defintion_main.sv'
filename = 'reset_defintion_main.sv'

##zhangqian
#ws.delete_cols(8,3)


rows=[]
cols=[]

#tup = ('func','out_sig','hw','sw')

#处理表格时暂不考虑第一列
#for row in ws.iter_rows(min_col = 0, max_col = 13):
for row in ws.iter_rows(min_col = 0, max_col = 9):
    rows.append(row)

#处理表格时不考虑第一行 
for col in ws.iter_cols(min_row = 1, max_row = 14):
    cols.append(col)


#zhangqian
#ws.delete_cols(8,3)

#define list name
reset_name_l = []
reset_num_l = []
reset_num_c = []
hw_reset_l = []
hw_reset_c = []
md_reset_l = []
md_reset_c = []
wdt_reset_c = []
wdt_reset_l = []
sw_reset_c = []
sw_reset_l = []
cpu_reset_c=[]
cpu_reset_l=[]
fsm_reset_c = []
fsm_reset_l = []
pd_reset_c=[]
pd_reset_l=[]

#定义字典
hw_reset_d = {}
md_reset_d = {}

###列操作
#reset name output list
with open(filename, 'w') as file_object:
    for elem in cols[1]:
        if(elem.value == None):
            continue
        else:
            reset_name_l.append(elem.value)
            #file_object.write(elem.value + '\n')
#删除第一行内容
reset_name_l.pop(0)

#pprint.pprint(reset_name_l)
#print("reset_name_l = {}".format(reset_name_l))

#reset Number list
with open(filename, 'a+') as file_object:
    for elem in cols[0]:
        if(elem.value == None):
            continue
        else:
            #reset_num_c.append(reset_name_l[elem.row-2].replace(r"_n_o$|_n$",'_num_c'))
            reset_num_c.append(re.sub(r"_n_o$|_n$",'_num_c',reset_name_l[elem.row-2]))
            reset_num_l.append(elem.row-2)
            elem.value = elem.row-2
            #file_object.write(elem.value + '\n')
reset_num_l.pop(0)
reset_num_c.pop(0)
#pprint.pprint(reset_num_l)
#print("reset_num_l = {}".format(reset_num_l))
#print("reset_num_c = {}".format(reset_num_c))

#HW reset list
with open(filename, 'a+') as file_object:
    for elem in cols[2]:
        if(elem.value == 'U'):
            #hw_reset_c.append(reset_name_l[elem.row-2].replace(r'_n_o$|_n$','_hw_c'))
            hw_reset_c.append(re.sub(r"_n_o$|_n$",'_hw_c',reset_name_l[elem.row-2]))
            hw_reset_l.append(0)
            elem.value = hw_reset_c[elem.row-2]
        elif(elem.value == 'R'):
            hw_reset_c.append(re.sub(r"_n_o$|_n$",'_hw_c',reset_name_l[elem.row-2]))
            hw_reset_l.append(1)
            elem.value = hw_reset_c[elem.row-2]
        else:
            print(elem.value)
            continue
          
#pprint.pprint(hw_reset_l)
#print("hw_reset_c = {}".format(hw_reset_c))
print("hw_reset_l = {}".format(hw_reset_l))

#Module reset list
with open(filename, 'a+') as file_object:
    for elem in cols[3]:
        if(elem.value == 'U'):
            #md_reset_c.append(reset_name_l[elem.row-2].replace(r'_n_o$|_n$','_md_c'))
            md_reset_c.append(re.sub(r"_n_o$|_n$",'_md_c',reset_name_l[elem.row-2]))
            md_reset_l.append(0)
            elem.value = md_reset_c[elem.row-2]
        elif(elem.value == 'R'):
            md_reset_c.append(re.sub(r"_n_o$|_n$",'_md_c',reset_name_l[elem.row-2]))
            md_reset_l.append(1)
            elem.value = md_reset_c[elem.row-2]
        else:
            continue

#print("md_reset_l = {}".format(md_reset_l))

#WDT reset list
with open(filename, 'a+') as file_object:
    for elem in cols[4]:
        if(elem.value == 'R'):
            #wdt_reset_c.append(reset_name_l[elem.row-2].replace(r'_n_o$|_n$','_wdt_c'))
            wdt_reset_c.append(re.sub(r"_n_o$|_n$",'_wdt_c',reset_name_l[elem.row-2]))
            wdt_reset_l.append(1)
            elem.value = wdt_reset_c[elem.row-2]
        elif(elem.value == 'U'):
            wdt_reset_c.append(re.sub(r"_n_o$|_n$",'_wdt_c',reset_name_l[elem.row-2]))
            wdt_reset_l.append(0)
            elem.value = wdt_reset_c[elem.row-2]
        else:
            continue

print("wdt_reset_l = {}".format(wdt_reset_l))
#print("wdt_reset_c = {}".format(wdt_reset_c))

#SW reset list
with open(filename, 'a+') as file_object:
    for elem in cols[5]:
        if(elem.value == 'R'):
            #sw_reset_c.append(reset_name_l[elem.row-2].replace(r'_n_o$|_n$','_sw_c'))
            sw_reset_c.append(re.sub(r"_n_o$|_n$",'_sw_c',reset_name_l[elem.row-2]))
            sw_reset_l.append(1)
            elem.value = sw_reset_c[elem.row-2]   
        elif(elem.value == 'U'):
            sw_reset_c.append(re.sub(r"_n_o$|_n$",'_sw_c',reset_name_l[elem.row-2]))
            sw_reset_l.append(0)
            elem.value = sw_reset_c[elem.row-2]
        else:
            continue
print("sw_reset_l = {}".format(sw_reset_l))

#cpu_reset list
with open(filename, 'a+') as file_object:
    #for elem in cols[11]:
    for elem in cols[6]:
        if(elem.value == 'R'):
            cpu_reset_c.append(re.sub(r"_n_o$|_n$",'_cpu_cnf',reset_name_l[elem.row-2]))
            cpu_reset_l.append(1)
            elem.value = cpu_reset_c[elem.row-2]   
        elif(elem.value == 'U'):
            cpu_reset_c.append(re.sub(r"_n_o$|_n$",'_cpu_cnf',reset_name_l[elem.row-2]))
            cpu_reset_l.append(0)
            elem.value = cpu_reset_c[elem.row-2]
        else:
            continue
print("cpu_reset_l= {}".format(cpu_reset_l))

#FSM reset list
with open(filename, 'a+') as file_object:
    for elem in cols[7]:
        if(elem.value == 'Pre'):
            #fsm_reset_c.append(reset_name_l[elem.row-2].replace(r'_n_o$|_n$','_fsm_c'))
            fsm_reset_c.append(re.sub(r"_n_o$|_n$",'_fsm_c',reset_name_l[elem.row-2]))
            fsm_reset_l.append(1)
            elem.value = fsm_reset_c[elem.row-2]
        elif(elem.value == 'FSM_reset'):
            print(elem.value)
            continue
        else:
            fsm_reset_c.append(re.sub(r"_n_o$|_n$",'_fsm_c',reset_name_l[elem.row-2]))
            fsm_reset_l.append(0)
            elem.value = fsm_reset_c[elem.row-2]
##fsm_reset_l.pop(0)
##fsm_reset_c.pop(0)

print("fsm_reset_l = {}".format(fsm_reset_l))
#print("fsm_reset_c = {}".format(fsm_reset_c))


#pd_reset list
with open(filename, 'a+') as file_object:
    #for elem in cols[10]:
    for elem in cols[8]:
        if(elem.value == 'R'):
            pd_reset_c.append(re.sub(r"_n_o$|_n$",'_pd_cnf',reset_name_l[elem.row-2]))
            pd_reset_l.append(1)
            elem.value = pd_reset_c[elem.row-2]   
        elif(elem.value == 'U'):
            pd_reset_c.append(re.sub(r"_n_o$|_n$",'_pd_cnf',reset_name_l[elem.row-2]))
            pd_reset_l.append(0)
            elem.value = pd_reset_c[elem.row-2]
        else:
            continue
print("pd_reset_l= {}".format(pd_reset_l))

#part two 行操作
#字符串直接写入
#os.linesep代表当前操作系统上的换行符
with open(filename, 'a+') as file_object:
    #file_object.writelines("package" + " reset_definition;" + os.linesep)
    #reset number 定义
    file_object.writelines("parameter" + " reset_num_main = "  + str(max(reset_num_l)+1) + ";" + os.linesep)
    #parameter定义
    for index in range(len(reset_num_l)):
        file_object.writelines("parameter " + str(reset_num_c[index]) + "  = " + "8'd" + str(reset_num_l[index]) + ";" + os.linesep)
        file_object.writelines("parameter " + str(hw_reset_c[index]) + "  = " + "1'b" + str(hw_reset_l[index]) + ";" + os.linesep)
        file_object.writelines("parameter " + str(md_reset_c[index]) + "  = " + "1'b" + str(md_reset_l[index]) + ";" + os.linesep)
        file_object.writelines("parameter " + str(wdt_reset_c[index]) + " = " + "1'b" + str(wdt_reset_l[index]) + ";" + os.linesep)
        file_object.writelines("parameter " + str(sw_reset_c[index]) + " = " + "1'b" + str(sw_reset_l[index]) + ";" + os.linesep)
        file_object.writelines("parameter " + str(cpu_reset_c[index]) + " = " + "1'b" + str(cpu_reset_l[index]) + ";" + os.linesep)
        file_object.writelines("parameter " + str(fsm_reset_c[index]) + " = " + "1'b" + str(fsm_reset_l[index]) + ";" + os.linesep)
        file_object.writelines("parameter " + str(pd_reset_c[index]) + " = " + "1'b" + str(pd_reset_l[index]) + ";" + os.linesep)

    #数组定义
    file_object.writelines(os.linesep)
    file_object.writelines("typedef struct packed {  "  + os.linesep)
    file_object.writelines("integer                  reset_index;"  + os.linesep)
    file_object.writelines("logic                    hw_rst_cnf;"  + os.linesep)
    file_object.writelines("logic                    mod_rst_req;"  + os.linesep)
    file_object.writelines("logic                    wdt_rst_req;"  + os.linesep)
    file_object.writelines("logic                    sw_rst_cnf;"  + os.linesep)
    file_object.writelines("logic                    cpu_rst_cnf;"  + os.linesep)
    file_object.writelines("logic                    rst_rls_ctrl;"  + os.linesep)
    file_object.writelines("logic                    pd_rst_cnf;"  + os.linesep)
    file_object.writelines("}scu_reset_define_main;"  + os.linesep)
    file_object.writelines(os.linesep)

    #数组赋值
    file_object.writelines("const scu_reset_define_main main_reset_matrix" + " [0: " + str(max(reset_num_l)) + "]" + "= '{" + os.linesep)

    print(len(reset_num_l))
    for index in range(len(reset_num_l)):
        print(index)
        module_name = str(rows[index][0].value)
        print(module_name)
        #print(rows[index][0])
        list = []
        for elem in rows[index+1]:
            list.append(str(elem.value))
        print(",   ".join(list))
        #删除reset_num_l 的信息，只保留constant
        list_pop = list.pop(1)
        #替换列表的 [] 为 {}
        str_reset = str(",   ".join(list))
        str_reset = str_reset.replace("[","{")
        str_reset = str_reset.replace("]","}")
        #写入sv 文件
        if index != max(reset_num_l):
            file_object.writelines("'{" + str_reset + "}," + os.linesep)
        else: 
            file_object.writelines("'{" + str_reset + "}};" + os.linesep)
    
   # file_object.writelines("endpackage" + os.linesep)


wb.save('test_scu.xlsx')
